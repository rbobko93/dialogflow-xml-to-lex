import json
import zipfile

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet import Worksheet


def get_file_name():
    print('File containing Dialogflow intents has to be in the same folder.')

    while True:
        try:
            file_name = input('What is the file name?')
            if len(file_name) == 0:
                print('File name cannot be empty.')

            if len(file_name) > 0:
                break
        except Exception:
            print('Oops! Something went wrong. Please try again.')

    return file_name


def get_prefix():
    print('You can also set prefix for new intent name. If you do not want any prefix enter empty value')

    return input('What is the intent name prefix?')


def get_sheet_number():
    while True:
        try:
            sheet_number = int(input('Starting with 1, what sheet number should be used?'))

            if sheet_number is None:
                print('Value cannot be empty.')
            else:
                break
        except Exception as err:
            print('Oops. Something went wrong. Remember to give me a number.')
            print('Error message: ', err)

    return sheet_number - 1


def load_file(file_name):
    while True:
        try:
            workbook = load_workbook(file_name)
            print('File loaded')
            return workbook
        except Exception as err:
            print('Oops! Could not load the file.')
            print('Error message: ', err)
            file_name = get_file_name()


def load_sheet(workbook: Workbook, sheet_number: int):
    try:
        worksheet = workbook[workbook.sheetnames[sheet_number]]
        print('Sheet retrieved')
        return worksheet
    except Exception as err:
        print('Oops! Could not load sheet from file.')
        print('Error message: ', err)


def convert_intents_to_lex(sheet: Worksheet, intent_name_prefix: str, zip_file):
    intent_object = {}

    for i in range(1, sheet.max_row):
        intent_name_cell = sheet.cell(i + 1, 1)
        utterance_cell = sheet.cell(i + 1, 2)
        response_cell = sheet.cell(i + 1, 3)

        if intent_name_cell.value:
            if 'metadata' not in intent_object:
                intent_object = create_new_lex_intent_object(intent_name_cell.value, intent_name_prefix)

            if utterance_cell.value:
                intent_object['sampleUtterances'].append(utterance_cell.value)

            if response_cell.value:
                message = {
                    "contentType": "PlainText",
                    "content": response_cell.value
                }
                intent_object['conclusionStatement']['messages'].append(message)
        else:
            if 'metadata' in intent_object:
                save_intent_to_file(intent_object, zip_file)
            intent_object = {}


def create_new_lex_intent_object(intent_name: str, prefix: str):
    intent_object = get_empty_lex_intent_object()
    intent_name = prefix + parse_name(intent_name)
    resource = {
        "name": intent_name,
        "version": 1,
        "fulfillmentActivity": {
            "type": "ReturnIntent"
        }
    }
    intent_object['resource'] = resource

    return intent_object


def get_empty_lex_intent_object():
    return {
        "metadata": {
            "schemaVersion": "1.0",
            "importType": "LEX",
            "importFormat": "JSON"
        },
        "resource": {},
        "sampleUtterances": [],
        "slots": [],
        "conclusionStatement": {
            "messages": []
        },
        "slotTypes": []
    }


def save_intent_to_file(intent_object, zip_file):
    file_name = intent_object['resource']['name'] + '.json'
    zip_file.writestr(file_name, json.dumps(intent_object, indent=4))
    print(file_name + ' saved in ' + zip_file.filename)


def parse_name(name: str):
    split = name.split('.')
    intent_name = split[-1]
    intent_category = split[-2].capitalize()
    words = intent_name.split("_")
    result = intent_category + '_'
    for w in words:
        result += w.capitalize()

    return result


def create_zip_file():
    try:
        zip_file = zipfile.ZipFile('intents.zip', 'w')
    except Exception as err:
        print('Oops! Something went wrong while creating zip file.')

    return zip_file


def main():
    print('Starting Dialogflow to Lex conversion.')

    file_name = get_file_name()
    intent_name_prefix = get_prefix()
    sheet_number = get_sheet_number()
    workbook = load_file(file_name)
    worksheet = load_sheet(workbook, sheet_number)
    zip_file = create_zip_file()
    convert_intents_to_lex(worksheet, intent_name_prefix, zip_file)

    print('Finished converting Dialogflow to Lex intents')


main()
